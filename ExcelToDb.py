import sys
import subprocess
import openpyxl
import re

## ソース出力フォルダー
SRC_OUTPUT_DIRECTORY = '../../'

#DB設計書の型：MySqlの型
mysqlTypeConv = {'tinyint':'TINYINT', 'bigint':'BIGINT', 'integer':'INTEGER', 'Integer':'INTEGER', 'varchar(n)':'STRING', 'boolean':'BOOLEAN', 'timestamp':'DATE', 'date':'DATE', 'time':'TIME', 'datetime': 'DATE'}
#DB設計書の型：node.jsの型
nodeTypeConv = {'tinyint':'number', 'bigint':'number', 'integer':'number', 'Integer':'number', 'varchar(n)':'string', 'boolean':'boolean', 'timestamp':'Date', 'date':'Date', 'time':'Date', 'datetime': 'Date'}

def snake_to_lower_camel(text):
    return re.sub('_([a-zA-Z0-9])', lambda m: m.group(1).upper(), text)

def snake_to_upper_camel(text):
    return re.sub('_([a-zA-Z0-9])', lambda m: m.group(1).upper(), text.capitalize())

def top_lower(text):
    return text[0].lower() + text[1:]

def top_upper(text):
    return text[0].upper() + text[1:]

################
# メイン処理
################

## コマンドパラメータ取得
args = sys.argv
if 2 != len(args):
    print('invalid args')
    print('usage:')
    print("> python " + args[0] + " path/to/DB設計.xlsx")
    sys.exit("Error")

## ワークブックOPEN
wb = openpyxl.load_workbook(args[1])
ws = wb.sheetnames

tsvfile = open('db.tsv', 'w')

tables = {}
uniqueKey = {}
fKey = {}
hasMany = {}
belongsTo = {}

#自動生成できないリレーションを追加
hasMany['control_item_api'] = []
hasMany['control_item_api'].append('\n'.join([
    "    models.controlItemApi.hasMany(models.roleControlItem, {",
    "      onUpdate: 'CASCADE',",
    "      foreignKey: 'controlItemId',",
    "      targetKey: 'id'",
    "    });\n"
]))
hasMany['role_control_item'] = []
hasMany['role_control_item'].append('\n'.join([
    "    models.roleControlItem.hasMany(models.controlItemApi, {",
    "      onUpdate: 'CASCADE',",
    "      foreignKey: 'controlItemId',",
    "      targetKey: 'id'",
    "    });\n"
]))

#作成しないシート名
skipSheets = [
'CSVアップロード履歴詳細'
]

tt = {}
## ワークシート順次処理
for shname in wb.sheetnames:
    sheet = wb[shname]
    if 'テーブル名(論理)' in str(sheet.cell(row=4, column=1).value) and (shname not in skipSheets):
        #print(sheet.cell(row=4, column=11).value)
        tableName = sheet.cell(row=5, column=11).value
        tables[tableName] = []
        uniqueKey[tableName] = []
        uniqueKeyFlg = False
        fKey[tableName] = []
        fKeyFlg = False
        colFlg = True
        if tableName not in belongsTo:
            belongsTo[tableName] = []

        maxRow = wb[shname].max_row + 1
        for i in range(10, maxRow):

            if sheet.cell(row=i, column=1).value == '複合一意キー':
                uniqueKeyFlg = True
                continue
            if sheet.cell(row=i, column=1).value == '外部キー':
                fKeyFlg = True
                continue

            if not sheet.cell(row=i, column=1).value:
                uniqueKeyFlg = False
                colFlg = False
                fKeyFlg = False
                continue

            if uniqueKeyFlg and sheet.cell(row=i, column=1).value != 'なし':
                uniqueKey[tableName].append(sheet.cell(row=i, column=1).value)
                continue

            if fKeyFlg:
                    belongsTo[tableName].append('\n'.join([
                        "    models." + snake_to_lower_camel(tableName) + ".belongsTo(models." + snake_to_lower_camel(refTable) + ", {",
                        "      onUpdate: 'CASCADE',",
                        "      foreignKey: '" + (foreignKey) + "',",
                        "      targetKey: '" + (sourceKey) + "'",
                        "    });\n"
                    ]))
                continue

            if not colFlg:
                continue

            colmunName = snake_to_lower_camel(sheet.cell(row=i, column=11).value)
           
            autoIncrement = sheet.cell(row=i, column=53).value
            defaultValue = sheet.cell(row=i, column=59).value
            name = sheet.cell(row=i, column=1).value
            tables[tableName].append([
                                      colmunName #0
                                    , colmunType #1
                                    , primaryKey #2
                                    , autoIncrement #3
                                    , notNull #4
                                    , field #5
                                    , defaultValue #6
                                    , name #7
                                    , colmunLength #8
                                    ])
            if colmunType not in tt:
                tt[colmunType] = 0
            tt[colmunType] += 1

            tsvfile.write(sheet.cell(row=4, column=11).value)
tsvfile.close()

dbInterfaceFile = open(SRC_OUTPUT_DIRECTORY)
importList = ['import { Sequelize } from \'sequelize\';\n']
interfaceList = ['  sequelize: Sequelize;\n']

for tableName in tables.keys():
    importList.append('import { ' + snake_to_upper_camel(tableName) + 'Model } from \'./tables/' + snake_to_lower_camel(tableName) + '\';\n')
    interfaceList.append('  ' + snake_to_lower_camel(tableName) + ': ' + snake_to_upper_camel(tableName) + 'Model;\n')
    indexList.append('import { create' + snake_to_upper_camel(tableName) + 'Model } from \'./tables/' + snake_to_lower_camel(tableName) + '\';\n')
    exportList.append('  ' + snake_to_lower_camel(tableName) + ': create' + snake_to_upper_camel(tableName) + 'Model(sequelize)')


dbInterfaceFile.write(''.join(importList))
#####
indexFile.write(',\n'.join(exportList) + '\n')
#####
for tableName, v in tables.items():

    file = open(snake_to_lower_camel(tableName) + '.ts', 'w')
    daoFile = open(snake_to_lower_camel(tableName) + '.ts', 'w')
    if tableName in hasMany:
            file.write('\n'.join(hasMany[tableName]))
            file.write('\n')
        if tableName in belongsTo:
            file.write('\n'.join(belongsTo[tableName]))
            file.write('\n')
        file.write('  };\n')
    file.write(',\n\n'.join(column) + '\n')
    file.write('    },\n')
    file.write('    {\n')
    file.write('      sequelize,\n')
    file.write('      freezeTableName: true,\n')
    file.write('      timestamps: false,\n')
    file.write('      tableName: \'' + tableName + '\',\n')
    file.write('      modelName: \''+ snake_to_lower_camel(tableName) + '\'\n')
    file.write('    }\n')

    daoFile.write('import { db } from \'../mysqlModels\';\n')
    daoFile.write('import { Transaction } from \'sequelize\';\n')
    daoFile.write('  }\n')
    daoFile.write('}\n')
    daoFile.close()
print('end!')
